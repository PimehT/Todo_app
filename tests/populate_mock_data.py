#!/usr/bin/env python3
"""
Run this script to clear the database, and then populate it with data
"""

FILE_PATH = "./mock_data.xlsx"


# DELETE EXISTING DATABASE
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.sql import text


# Define the connection string
username = 'root'
password = 'password'
host = 'localhost'

# Create the engine to connect to the MySQL server
engine = create_engine(f'mysql+mysqldb://{username}:{password}@{host}',
                      pool_pre_ping=True)

# Connect to the engine
connection = engine.connect()

SQL_STATEMENTS = [
"DROP DATABASE IF EXISTS todo_test_db;",
"CREATE DATABASE IF NOT EXISTS todo_test_db;",
"CREATE USER IF NOT EXISTS 'todo_test'@'localhost' IDENTIFIED BY 'todo_test_pwd';",
"GRANT ALL PRIVILEGES ON `todo_test_db`.* TO 'todo_test'@'localhost';",
"GRANT SELECT ON `performance_schema`.* TO 'todo_test'@'localhost';",
"FLUSH PRIVILEGES;",
]

try:
    # Execute SQL statements
    for stmt in SQL_STATEMENTS:
        connection.execute(text(stmt))

except SQLAlchemyError as e:
    print(f"Error: {e}")

finally:
    # Close the connection
    connection.close()
    print("DATABASE reset successfully.")

import sys
sys.path.append('..')
# POPULATE DATABASE
import openpyxl as op
import models
from models.base_model import BaseModel, datetime_format
from models.category import Category
from models.comment import Comment
from models.task import Task, task_category
from models.user import User
from sqlalchemy import insert
from datetime import datetime

data_constraints = {
    # 'sheet_name': ( class, [{attr: type, ...}] )
    'users': (User,),
    'tasks': (Task,),
    'categories': (Category,),
    'comments': (Comment, ),
}

association_tables = {
    'task_category': task_category,
}


wb = op.load_workbook(FILE_PATH, data_only=True)

entries = 0

def get_value(ws, row, col):
    return ws.cell(row, col).value

data_log = dict()

def worksheet(ws_name, obj, types=dict()):
    global entries
    ws = wb[ws_name]
    # map columns
    col_map = dict()
    for col in range(1, ws.max_column+1, 1):
        v = ws.cell(1, col).value
        if not v or len(v) == 0:
            break
        col_map[v] = col

    for row in range(2, ws.max_row+1, 1):
        attr_dict = dict()
        if ws.cell(row, 1).value is None or len(str(ws.cell(row, 1).value)) == 0:
            break;
        for attr, col in col_map.items():
            if attr in ['updated_at', 'created_at']:
                continue # skip
            if attr == 'password':
                value = str(get_value(ws, row, col))
                # value = generate_password_hash(get_value(ws, row, col),
                #                                method="pbkdf2:sha1:1000",
                #                                salt_length=8)
            elif attr in ['deadline']:
                value = datetime.strptime(get_value(ws, row, col), datetime_format)
            elif attr in types:
                value = types[attr](get_value(ws, row, col))
            else:
                value = str(get_value(ws, row, col))
            attr_dict[attr] = value

        instance = obj(**attr_dict)
        instance.save()
        entries += 1
        obj_name = obj._sa_class_manager.class_.__name__
        data_log[f"{obj_name}.{instance.id}"] = instance

"""Populate tables."""
for ws_name, data in data_constraints.items():
    class_ = data[0]
    if len(data) > 1:
        types = data[1]
    else:
        types = dict()
    worksheet(ws_name, class_, types)

"""Association tables."""
for table, table_obj in association_tables.items():
    ws = wb[table]
    # map columns
    col_map = dict()
    for col in range(1, ws.max_column+1, 1):
        v = ws.cell(1, col).value
        if not v or len(v) == 0:
            break
        col_map[v] = col
    values_list = list()
    # loop through worksheet
    for row in range(2, ws.max_row+1, 1):
        attr_dict = dict()
        for attr, col in col_map.items():
            value = ws.cell(row, col).value
            attr_dict[attr] = value
        # add values to list
        values_list.append(attr_dict)
        entries += 1

    models.storage.close()
    # insert into table
    with models.storage._DBStorage__engine.connect() as conn:
        result = conn.execute(insert(table_obj), values_list,)
        conn.commit()

print('DONE!\tEntries: ', entries)
# print(data_log)
